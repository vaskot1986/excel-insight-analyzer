import pandas as pd
import openpyxl
import random

def load_and_filter_excel(file_path):
    """
    Load the Excel file and filter out rows where 'Ignore' is marked as 'yes'.
    """
    # Load the Excel file
    df = pd.read_excel(file_path)
    
    # Print the columns to debug
    print("Columns in the loaded DataFrame:", df.columns.tolist())
    
    # Filter out rows where 'Ignore' is marked as 'yes'
    filtered_df = df[df['Ignore'].str.lower() != 'yes']
    
    return filtered_df

def analyze_sheet_structure_by_marker(file_path, sheet_name, marker):
    """
    Analyze the structure of the specified sheet and identify table columns based on an exact match of the marker in the cell.
    """
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    # Normalize the marker for comparison
    normalized_marker = marker.strip().lower()
    
    # Find the row containing the marker
    header_row_idx = None
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if any(cell is not None and str(cell).strip().lower() == normalized_marker for cell in row):
            header_row_idx = i
            print("Header line found at row:", header_row_idx + 1)  # openpyxl is 1-indexed
            break
    
    if header_row_idx is None:
        raise ValueError(f"Marker '{marker}' not found in any cell of sheet '{sheet_name}' in file '{file_path}'")
    
    # Extract the column headers based on the identified header row
    header_row = list(sheet.iter_rows(min_row=header_row_idx+1, max_row=header_row_idx+1, values_only=True))[0]
    
    # Limit the number of consecutive empty values printed to 5
    filtered_header_row = []
    empty_count = 0
    for cell in header_row:
        if cell is None:
            empty_count += 1
            if empty_count <= 5:
                filtered_header_row.append(cell)
        else:
            empty_count = 0
            filtered_header_row.append(cell)
    
    print("Filtered header row:", filtered_header_row)
    
    return filtered_header_row, header_row_idx

def remove_trailing_none(headers):
    """
    Remove trailing None values from the list of headers.
    """
    while headers and headers[-1] is None:
        headers.pop()
    return headers

def gather_column_info(df, headers):
    """
    Gather information about each column including data type, examples, and if the column is always empty.
    """
    column_info = {}
    sample_size = 50
    
    for column in headers:
        if column in df.columns:
            data = df[column].head(sample_size).dropna()
            data_type = data.dtype
            examples = data.sample(min(len(data), 2)).tolist()  # Extracting examples
            print(f"Column: {column}, Data Type: {data_type}, Examples: {examples}")  # Debug line
            is_always_empty = len(data) == 0
            
            column_info[column] = {
                "Data Type": str(data_type),
                "Examples": examples,
                "Always Empty": is_always_empty,
            }
            print(f"Column Info for {column}: {column_info[column]}")  # Added debug line
        else:
            column_info[column] = {
                "Data Type": "N/A",
                "Examples": [],
                "Always Empty": True,
            }
            print(f"Column {column} not found in DataFrame")  # Added debug line
    
    return column_info


def save_column_info_to_excel(column_info_list, output_file):
    df = pd.DataFrame(column_info_list)
    print("DataFrame to be saved:\n", df)  # Added debug line to check DataFrame content
    df.to_excel(output_file, index=False)
    print(f"Column info saved to {output_file}")

